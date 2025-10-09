import os
import aubio
import numpy as np
import soundfile as sf
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from collections import defaultdict
import time
import psutil

# ================= CONFIG =====================
AUDIO_FOLDER = "./UsedFilesInResearchPaper/sound/real_sax_notes"
OUTPUT_XLSX = "new_voted_real_sax_notes.xlsx"

SAMPLERATE = 44100
WIN_SIZE = 4096
HOP_SIZE = 512

note_frequencies = {
    "C3": 130.81, "C#3": 138.59, "D3": 146.83, "D#3": 155.56,
    "E3": 164.81, "F3": 174.61, "F#3": 185.00, "G3": 196.00,
    "G#3": 207.65, "A3": 220.00, "A#3": 233.08, "B3": 246.94,
    "C4": 261.63, "C#4": 277.18, "D4": 293.66, "D#4": 311.13,
    "E4": 329.63, "F4": 349.23, "F#4": 369.99, "G4": 392.00,
    "G#4": 415.30, "A4": 440.00, "A#4": 466.16, "B4": 493.88,
    "C5": 523.25, "C#5": 554.37, "D5": 587.33, "D#5": 622.25,
    "E5": 659.25, "F5": 698.46, "F#5": 739.99, "G5": 783.99,
    "G#5": 830.61, "A5": 880.00
}


# ==============================================

# ---------------- ANALYSIS FUNCTIONS ----------------
def process_file_detailed(file_path):
    audio, sr = sf.read(file_path)
    if len(audio.shape) > 1:
        audio = np.mean(audio, axis=1)
    if audio.dtype != np.float32:
        audio = audio.astype(np.float32)
    max_val = np.max(np.abs(audio))
    if max_val > 0:
        audio = audio / max_val * 0.95

    detector_configs = {
        "yin": {"tolerance": 0.85, "silence": -40},
        "mcomb": {"tolerance": 0.85, "silence": -40},
        "schmitt": {"tolerance": 0.3, "silence": -40}
    }

    detectors = {}
    for name, cfg in detector_configs.items():
        d = aubio.pitch(name, WIN_SIZE, HOP_SIZE, SAMPLERATE)
        d.set_unit("Hz")
        d.set_silence(cfg["silence"])
        d.set_tolerance(cfg["tolerance"])
        detectors[name] = d

    performance_data = {}
    pitches_per_detector = defaultdict(list)

    for name, det in detectors.items():
        start_time = time.time()
        process = psutil.Process()
        cpu_before = process.cpu_percent()
        frame_count = 0
        pitches = []

        while frame_count + WIN_SIZE <= len(audio):
            frame = audio[frame_count:frame_count + HOP_SIZE]
            pitch = det(frame)[0]
            if pitch > 0:
                pitches.append(pitch)
            frame_count += HOP_SIZE

        end_time = time.time()
        cpu_after = process.cpu_percent()

        pitches_per_detector[name] = pitches
        performance_data[name] = {
            'processing_time': end_time - start_time,
            'cpu_usage': cpu_after - cpu_before,
            'frames_processed': len(pitches)
        }

    return pitches_per_detector, performance_data


def calculate_accuracy_metrics(pitches, ground_truth):
    if not pitches:
        return {'mean': 0, 'median': 0, 'std': 0,
                'abs_error': float('inf'),
                'rel_error_percent': float('inf'),
                'cents_error': float('inf')}

    skip = int(len(pitches) * 0.15)
    pitches = pitches[skip:] if len(pitches) > skip else pitches

    mean_pitch = np.mean(pitches)
    median_pitch = np.median(pitches)
    std_pitch = np.std(pitches)

    abs_error = abs(mean_pitch - ground_truth)
    rel_error = (abs_error / ground_truth) * 100 if ground_truth else float('inf')
    cents_error = 1200 * np.log2(mean_pitch / ground_truth) if (mean_pitch > 0 and ground_truth > 0) \
        else float('inf')

    return {
        'mean': mean_pitch,
        'median': median_pitch,
        'std': std_pitch,
        'abs_error': abs_error,
        'rel_error_percent': rel_error,
        'cents_error': abs(cents_error)
    }


# ---------------- EXCEL COLORS ----------------
from openpyxl.styles import PatternFill

COLOR_EXCELLENT = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
COLOR_GOOD = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
COLOR_POOR = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
COLOR_TERRIBLE = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")


def get_accuracy_color(cents_error):
    if cents_error <= 10:
        return COLOR_EXCELLENT
    elif cents_error <= 50:
        return COLOR_GOOD
    elif cents_error <= 100:
        return COLOR_POOR
    else:
        return COLOR_TERRIBLE


def get_voted_color(voted_pitch, ground_truth):
    """
    Dynamic coloring based on note range:
    - 0–440 Hz: 2 Hz tolerance (excellent)
    - 440–1000 Hz: 3 Hz tolerance (excellent)

    This approximately mimics the cent errors used in the base algorithms
    """
    if voted_pitch is None or ground_truth is None:
        return COLOR_TERRIBLE

    diff = abs(voted_pitch - ground_truth)

    # Set dynamic tolerances
    if ground_truth <= 440:
        excellent_tol = 2.0
        good_tol = 5.0
        poor_tol = 10.0
    else:
        excellent_tol = 3.0
        good_tol = 7.0
        poor_tol = 15.0

    if diff <= excellent_tol:
        return COLOR_EXCELLENT
    elif diff <= good_tol:
        return COLOR_GOOD
    elif diff <= poor_tol:
        return COLOR_POOR
    else:
        return COLOR_TERRIBLE


# ---------------- OCTAVE-AWARE VOTING ----------------
def fused_pitch_octave_aware(medians, stds):
    """
    Octave-aware voting with both upward and downward correction.
    1. Detect harmonic relationships (≈1/4×, 1/2×, 2×, 3×, 4×).
    2. Normalize all to closest “base” fundamental.
    3. Compute weighted mean using 1/std² after outlier filtering.
    """

    if not medians:
        return None

    freqs = np.array(list(medians.values()))
    names = list(medians.keys())
    std_array = np.array([stds.get(n, 1.0) for n in names])

    # Step 1: Normalize to the closest fundamental by detecting octave/harmonic jumps
    normalized = freqs.copy()
    corrected = False
    for i in range(len(freqs)):
        for j in range(len(freqs)):
            if i == j or corrected:
                continue
            ratio = freqs[i] / freqs[j]
            for mul in [0.25, 0.5, 2, 3, 4]:  # check both downwards and upwards
                if 0.95 * mul < ratio < 1.05 * mul:
                    normalized[i] = freqs[i] / mul
                    corrected = True
                    break

    # Step 2: Weighted mean by stability
    weights = 1 / (std_array ** 2 + 1e-6)  # added a small number to not divide with 0
    weighted_mean = np.sum(normalized * weights) / np.sum(weights)

    # Step 3: Outlier filtering
    # harsh filtering to not include strange results
    valid = np.abs(normalized - weighted_mean) < 0.05 * weighted_mean
    if np.sum(valid) >= 2:
        return np.mean(normalized[valid])
    return weighted_mean


# ---------------- MAIN FUNCTION ----------------
def main():
    if os.path.exists(OUTPUT_XLSX):
        os.remove(OUTPUT_XLSX)
        print(f"🗑️ Removed existing file: {OUTPUT_XLSX}")

    files_to_process = []
    for note, freq in note_frequencies.items():
        path = os.path.join(AUDIO_FOLDER, f"{note}.wav")
        if os.path.exists(path):
            files_to_process.append((freq, path, note))
        else:
            print(f"⚠ Missing: {note}.wav")

    files_to_process.sort(key=lambda x: x[0])

    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Results"

    headers = [
        "Note", "File", "Detector", "Mean Hz", "Median Hz", "Std Dev",
        "Ground Truth Hz", "Abs Error Hz", "Rel Error %", "Cents Error",
        "Processing Time (s)", "Frames Processed"
    ]
    ws.append(headers)
    for c in ws[1]: c.font = Font(bold=True)

    summary = defaultdict(lambda: defaultdict(list))

    for freq, path, note in files_to_process:
        print(f"Processing {note}...")
        pitches_per_detector, perf_data = process_file_detailed(path)
        medians, stds = {}, {}

        for det, pitches in pitches_per_detector.items():
            metrics = calculate_accuracy_metrics(pitches, freq)
            medians[det] = metrics['median']
            stds[det] = metrics['std']

            if metrics['abs_error'] != float('inf'):
                summary[det]['abs'].append(metrics['abs_error'])
                summary[det]['rel'].append(metrics['rel_error_percent'])
                summary[det]['cents'].append(metrics['cents_error'])
                summary[det]['time'].append(perf_data[det]['processing_time'])
                summary[det]['ok'].append(1)
            else:
                summary[det]['ok'].append(0)

            row = [
                note, os.path.basename(path), det,
                round(metrics['mean'], 4), round(metrics['median'], 4), round(metrics['std'], 4),
                round(freq, 4), round(metrics['abs_error'], 4),
                round(metrics['rel_error_percent'], 2), round(metrics['cents_error'], 1),
                round(perf_data[det]['processing_time'], 4),
                perf_data[det]['frames_processed']
            ]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.fill = get_accuracy_color(metrics['cents_error'])

        # --------- Octave-aware voting ----------
        voted_pitch = fused_pitch_octave_aware(medians, stds)
        if voted_pitch is not None:
            row_voted = [
                note, os.path.basename(path), "VOTED",
                round(voted_pitch, 4), round(voted_pitch, 4), "",
                round(freq, 4), "", "", "",
                "", ""
            ]
            ws.append(row_voted)
            for cell in ws[ws.max_row]:
                cell.fill = get_voted_color(voted_pitch, freq)

            print(f"VOTED pitch for {note}: {voted_pitch:.2f} Hz")

    wb.save(OUTPUT_XLSX)
    print(f"✅ Saved results to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
