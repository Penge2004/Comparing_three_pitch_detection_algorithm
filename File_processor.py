import os
import aubio
import numpy as np
import soundfile as sf
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import time
import psutil
from collections import defaultdict

# ================= CONFIG =====================

AUDIO_FOLDER = "./1sec_60harm_6Hz_100cent_0db"
OUTPUT_XLSX = "1sec_60harm_6Hz_100cent_0db_badSaxConditions.xlsx"

SAMPLERATE = 44100
WIN_SIZE = 4096
HOP_SIZE = 512

note_frequencies = {
    "C3": 130.81, "C#3": 138.59, "D3": 146.83, "D#3": 155.56,
    "E3": 164.81, "F3": 174.61, "F#3": 185.00, "G3": 196.00,
    "G#3": 207.65, "A3": 220.00, "A#3": 233.08, "B3": 246.94,
    "C4": 261.63, "C#4": 277.18, "D4": 293.66, "D#4": 311.13,
    "E4": 329.63, "F4": 349.23, "F#4": 369.99, "G4": 392.00,
    "G#4": 415.30, "A4": 440.00, "A#4": 466.16, "B4": 493.88,
    "C5": 523.25, "C#5": 554.37, "D5": 587.33, "D#5": 622.25,
    "E5": 659.25, "F5": 698.46, "F#5": 739.99, "G5": 783.99,
    "G#5": 830.61, "A5": 880.00
}

# ==============================================

def process_file_detailed(file_path):
    """Process a WAV file with detailed analysis and performance metrics."""
    audio, sr = sf.read(file_path)

    # Ensure mono audio
    if len(audio.shape) > 1:
        audio = np.mean(audio, axis=1)

    # Convert to float32 if needed
    if audio.dtype != np.float32:
        audio = audio.astype(np.float32)

    # Normalize audio (but be careful not to over-normalize)
    max_val = np.max(np.abs(audio))
    if max_val > 0:
        audio = audio / max_val * 0.95  # Leave some headroom

    # Setup detectors with different configurations - most optimal for each algorithm
    detector_configs = {
        "yin": {"tolerance": 0.85, "silence": -40},
        "mcomb": {"tolerance": 0.85, "silence": -40},
        "schmitt": {"tolerance": 0.3, "silence": -40}
    }

    detectors = {}
    for name, config in detector_configs.items():
        det = aubio.pitch(name, WIN_SIZE, HOP_SIZE, SAMPLERATE)
        det.set_unit("Hz")
        det.set_silence(config["silence"])
        det.set_tolerance(config["tolerance"])
        detectors[name] = det

    # Performance tracking
    performance_data = {}
    pitches_per_detector = defaultdict(list)



    for name, det in detectors.items():
        start_time = time.time()
        process = psutil.Process()
        cpu_before = process.cpu_percent()

        frame_count = 0
        pitches = []

        while frame_count + WIN_SIZE <= len(audio):
            frame = audio[frame_count:frame_count + HOP_SIZE]

            pitch = det(frame)[0]

            # We let any positive frequency pass -> transparency
            if pitch > 0:
                pitches.append(pitch)

            frame_count += HOP_SIZE

        end_time = time.time()
        cpu_after = process.cpu_percent()

        pitches_per_detector[name] = pitches

        performance_data[name] = {
            'processing_time': end_time - start_time,
            'cpu_usage': cpu_after - cpu_before,
            'frames_processed': len(pitches)
        }

    return pitches_per_detector, performance_data


def calculate_accuracy_metrics(pitches, ground_truth):
    """Calculate comprehensive accuracy metrics."""
    if not pitches:
        return {
            'mean': 0,
            'median': 0,
            'std': 0,
            'abs_error': float('inf'),
            'rel_error_percent': float('inf'),
            'cents_error': float('inf')
        }


    STABILIZATION_FRAMES_PERCENT = 0.15 # 15percent of the audio is being skipped - "warm-up"
    skip_count = int(len(pitches) * STABILIZATION_FRAMES_PERCENT)
    if len(pitches) > skip_count:
        pitches = pitches[skip_count:]

    mean_pitch = np.mean(pitches)
    median_pitch = np.median(pitches)
    std_pitch = np.std(pitches)
    abs_error = abs(mean_pitch - ground_truth)
    rel_error = (abs_error / ground_truth) * 100

    # Calculate error in cents (musical intervals)
    if mean_pitch > 0 and ground_truth > 0:
        cents_error = 1200 * np.log2(mean_pitch / ground_truth)
    else:
        cents_error = float('inf')

    return {
        'mean': mean_pitch,
        'median': median_pitch,
        'std': std_pitch,
        'abs_error': abs_error,
        'rel_error_percent': rel_error,
        'cents_error': abs(cents_error)
    }


# Excel styling
COLOR_EXCELLENT = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
COLOR_GOOD = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light yellow
COLOR_POOR = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light pink
COLOR_TERRIBLE = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Light red


def get_accuracy_color(cents_error):
    """Determine color based on accuracy (cents error)."""
    if cents_error <= 10:  # Excellent (within 10 cents)
        return COLOR_EXCELLENT
    elif cents_error <= 50:  # Good (within 50 cents)
        return COLOR_GOOD
    elif cents_error <= 100:  # Poor (within 100 cents)
        return COLOR_POOR
    else:  # Terrible (over 100 cents)
        return COLOR_TERRIBLE


def main():
    # Remove existing output file if it exists
    if os.path.exists(OUTPUT_XLSX):
        os.remove(OUTPUT_XLSX)
        print(f"🗑️ Removed existing file: {OUTPUT_XLSX}")

    # Prepare files in frequency order
    files_to_process = []
    for note_name, freq in note_frequencies.items():
        wav_file = f"{note_name}.wav"
        file_path = os.path.join(AUDIO_FOLDER, wav_file)
        if os.path.exists(file_path):
            files_to_process.append((freq, file_path, note_name))
        else:
            print(f"⚠ File not found: {wav_file}")

    files_to_process.sort(key=lambda x: x[0])

    # Create Excel workbook with multiple sheets
    wb = Workbook()

    # Main results sheet
    ws_main = wb.active
    ws_main.title = "Detailed Results"

    headers = [
        "Note", "File", "Detector", "Mean Hz", "Median Hz", "Std Dev",
        "Ground Truth Hz", "Abs Error Hz", "Rel Error %", "Cents Error",
        "Processing Time (s)", "Frames Processed"
    ]
    ws_main.append(headers)

    # Make headers bold
    for cell in ws_main[1]:
        cell.font = Font(bold=True)

    # Summary sheet
    ws_summary = wb.create_sheet("Algorithm Summary")
    summary_headers = [
        "Algorithm", "Avg Abs Error Hz", "Avg Rel Error %", "Avg Cents Error",
        "Avg Processing Time", "Success Rate %"
    ]
    ws_summary.append(summary_headers)
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)

    # Collect summary statistics
    summary_stats = defaultdict(lambda: defaultdict(list))

    for freq, file_path, note_name in files_to_process:
        print(f"Processing {note_name}...")

        pitches_per_detector,  performance_data = process_file_detailed(file_path)

        for detector_name in pitches_per_detector.keys():
            pitches = pitches_per_detector[detector_name]
            perf = performance_data[detector_name]

            metrics = calculate_accuracy_metrics(pitches, freq)

            # Store for summary
            if metrics['abs_error'] != float('inf'):
                summary_stats[detector_name]['abs_errors'].append(metrics['abs_error'])
                summary_stats[detector_name]['rel_errors'].append(metrics['rel_error_percent'])
                summary_stats[detector_name]['cents_errors'].append(metrics['cents_error'])
                summary_stats[detector_name]['times'].append(perf['processing_time'])
                summary_stats[detector_name]['successes'].append(1)
            else:
                summary_stats[detector_name]['successes'].append(0)

            row = [
                note_name,
                os.path.basename(file_path),
                detector_name,
                round(metrics['mean'], 4) if metrics['mean'] != 0 else "N/A",
                round(metrics['median'], 4) if metrics['median'] != 0 else "N/A",
                round(metrics['std'], 4) if metrics['std'] != 0 else "N/A",
                round(freq, 4),
                round(metrics['abs_error'], 4) if metrics['abs_error'] != float('inf') else "FAIL",
                round(metrics['rel_error_percent'], 2) if metrics['rel_error_percent'] != float('inf') else "FAIL",
                round(metrics['cents_error'], 1) if metrics['cents_error'] != float('inf') else "FAIL",
                round(perf['processing_time'], 4),
                perf['frames_processed']
            ]

            ws_main.append(row)

            # Color code based on accuracy
            fill_color = get_accuracy_color(metrics['cents_error'])
            for cell in ws_main[ws_main.max_row]:
                cell.fill = fill_color

    # Generate summary statistics
    for detector_name, stats in summary_stats.items():
        if stats['abs_errors']:
            avg_abs_error = np.mean(stats['abs_errors'])
            avg_rel_error = np.mean(stats['rel_errors'])
            avg_cents_error = np.mean(stats['cents_errors'])
            avg_time = np.mean(stats['times'])
            success_rate = (sum(stats['successes']) / len(stats['successes'])) * 100
        else:
            avg_abs_error = avg_rel_error = avg_cents_error =  avg_time = 0
            success_rate = 0

        summary_row = [
            detector_name,
            round(avg_abs_error, 4),
            round(avg_rel_error, 2),
            round(avg_cents_error, 1),
            round(avg_time, 4),
            round(success_rate, 1)
        ]
        ws_summary.append(summary_row)

        # Color code summary based on cents error
        fill_color = get_accuracy_color(avg_cents_error)
        for cell in ws_summary[ws_summary.max_row]:
            cell.fill = fill_color

    # Auto-adjust column widths
    for ws in [ws_main, ws_summary]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(OUTPUT_XLSX)
    print(f"✅ Analysis complete! Detailed results saved to {OUTPUT_XLSX}")
    print("\nColor coding:")
    print("🟢 Green: Excellent accuracy (≤10 cents error)")
    print("🟡 Yellow: Good accuracy (≤50 cents error)")
    print("🟠 Pink: Poor accuracy (≤100 cents error)")
    print("🔴 Red: Terrible accuracy (>100 cents error)")


if __name__ == "__main__":
    main()
